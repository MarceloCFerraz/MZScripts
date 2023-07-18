import openpyxl
from utils.common import printLine, cprint

emails_index = 1
channels_index = 2


def channelNameColumnIndex(sheet):
    print("Searching for the 'Channel Name' column index")

    for column_index in range(1, sheet.max_column+1):
        cell = sheet.cell(row=1, column=column_index)
        value = cell.value

        if "Channel Name" == value:
            print("'Channel Name' found at column {}".format(column_index))
            printLine()
            channels_index = column_index
            return column_index
    
    print("COLUMN NOT FOUND!")
    printLine()
    return None
    

def emailsColumnIndex(sheet):
    print("Searching for the 'User E-Mail' column index")

    for column_index in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=column_index)
        value = cell.value
        print(value)

        if "User E-Mail" == value:
            print("'User E-Mail' found at column {}".format(column_index))
            printLine()
            emails_index = column_index
            return column_index
    
    print("COLUMN NOT FOUND!")
    printLine()
    return None


def emailsListFromChannelName(sheet, channel_key, channel_row_index):
    emails_list = []

    for row_index in range(channel_row_index, sheet.max_row):
        # searches for e-mails starting on the row the channel was located first (e.g. line 2)
        channel_name = sheet.cell(row=row_index, column=channels_index).value
        email = sheet.cell(row=row_index, column=emails_index).value
        
        if str(channel_name).strip() == str(channel_key).strip():
            emails_list.append(email)

    cprint(
        "Emails list for {} filled successfully with ".format(channel_key) +
        "{} items".format(len(emails_list)),
        "light_blue"
    )
    printListItems(emails_list)
    printLine()

    return emails_list


def printListItems(list):
    for item in list:
        print("-> {}".format(item))


def channelsDict(sheet):
    # channels_dict = 
    # {
    #   "channel name": [ "email 1 ", "email 2", ... ] 
    # }

    channels_dict = {}
    channels_list = []

    print("Starting filling the channels dictionary")
    printLine()

    for row_index in range(2, sheet.max_row):
        # it jumps the first line because it's where the header is located

        channel_name = sheet.cell(row=row_index, column=channels_index).value
        email = sheet.cell(row=row_index, column=emails_index).value
        
        if channel_name != None and email != None:
            channel_name = str(channel_name).strip()
            email = str(email).strip()
            if channel_name not in channels_list:
                print("Getting E-Mails list for Channel {}".format(channel_name))
                channels_list.append(channel_name)
                channels_dict[channel_name] = emailsListFromChannelName(sheet, channel_name, row_index)
    
    print("Finished filling the dictionary")
    printLine()
    return channels_dict
